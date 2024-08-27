import openpyxl

class ReadExcelData(object):

    def __init__(self, file_path: str, data_only: bool = True):
        """[summary]
        
        [description]
        
        Args:
            file_path (str): [文件地址]
            data_only (bool): [空值不要读取] (default: `True`)
        """

        self.file_path = file_path
        self.data_only = data_only

    def open_file(self):
        try:
            book = openpyxl.load_workbook(self.file_path, data_only=self.data_only)
        except Exception as e:
            raise
        return book

    def get_sheet(self):
        book = self.open_file()
        sheet_list = book.sheetnames
        return sheet_list

    def open_sheet(self, sheetname: str):
        book = self.open_file()
        sheet = book[sheetname]
        return sheet

    def get_data(self, sheetname: str):
        """[summary]
        
        [description]
        
        Args:
            sheetname (str): [读取sheet页名称]
        """
        sheet = self.open_sheet(sheetname)
        data = [row for row in sheet.iter_rows(values_only=self.data_only)]
        return data

    def get_header(self, sheetname: str, header_row=0):
        """[summary]
        
        [description]
        
        Args:
            sheetname (str): [读取sheet页]
            headers ([type]): [description]
            header_row (number): [第几行] (default: `0`)
        """

        data = self.get_data(sheetname)
        data_header = data[header_row]

        return data_header

    def get_cell_value(self, sheetname, row, col):
        """[summary]
        
        [description]
        
        Args:
            sheetname ([type]): [读取的sheet页]
            row ([type]): [行数]
            col ([type]): [列数]
        """
        sheet = self.open_sheet(sheetname)
        cell_value = sheet.cell(row=row, col=col)
        return cell_value

    def get_mergecell(self, sheetname: str):
        """[summary]
        
        [description]
        
        Args:
            sheetname (str): [读取的sheet页]
        """
        mergecell_list = []
        sheet = self.open_sheet(sheetname)
        merge_range = sheet.merged_cells.ranges
        for cell in merge_range:
            start_row = cell.min_row
            end_row = cell.max_row
            start_col = cell.min_col
            end_col = cell.max_col
            cell_value = sheet.cell(row=start_row, column=start_col).value

            # mergecell_tuple【起始行，结束行，起始列，结束列，】
            # mergecell_tuple = tuple([start_row, end_row, start_col, end_col, cell_value])
            mergecell_tuple = dict(start_row=start_row, end_row=end_row, start_col=start_col, end_col=end_col, cell_value=cell_value)

            mergecell_list.append(mergecell_tuple)
        return mergecell_list


if __name__ == '__main__':
    file_path = r'D:\我的坚果云\work\Sincetimes_DOW\2023-09-21 XCY DYL DOW damo 活动历史对比\damo活动\【日本】霸王天下-活动表2020年.xlsx'

    test = ReadExcelData(file_path)
    res = test.get_mergecell('活动排期')
    print(res)